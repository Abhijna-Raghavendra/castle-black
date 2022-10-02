from flask import Flask, render_template, request, jsonify, session
from flask_wtf import FlaskForm
from wtforms import StringField, PasswordField, SubmitField  
from wtforms.validators import DataRequired
from openpyxl.workbook import Workbook
from openpyxl import load_workbook
from werkzeug.security import generate_password_hash, check_password_hash
from dotenv import load_dotenv
from datetime import datetime, timedelta
from functools import wraps
from gpiozero import Servo
import jwt
import os
import RPi.GPIO as GPIO

app = Flask(__name__)
app.config['SECRET_KEY'] = os.environ.get("MDG_SECRET_KEY", "")

GPIO.setwarnings(False)
GPIO.setmode(GPIO.BCM)
GPIO.setup(4, GPIO.OUT)
servo = Servo(4)

wb = load_workbook('sheet.xlsx')
ws = wb['passwords']
logs = wb['logs']


class LoginForm(FlaskForm):
	uname = StringField("Username:", validators=[DataRequired()])
	pswd = PasswordField("Password:", validators=[DataRequired()])
	submit = SubmitField("Submit")
	def validate_on_submit(self):
		if self.uname.data:
			for col in ws.iter_cols(min_row=2, min_col=1, max_col=1):
				for cell in col:
					if cell.value == self.uname.data:
						return check_password_hash(ws['B' + str(cell.row)].value, self.pswd.data)
			return False
		else:
			return True

class PswdForm(FlaskForm):
	pswd1 = PasswordField("New password:", validators=[DataRequired()])
	pswd2 = PasswordField("Confirm Password:", validators=[DataRequired()])
	submit = SubmitField("Submit")
	def change_pswd(self, usr):
		if (self.pswd1.data == self.pswd2.data) and (self.pswd1.data != None):
			for col in ws.iter_cols(min_row=2, min_col=1, max_col=1):
				for cell in col:
					if cell.value == usr:
						ws['B' + str(cell.row)] = generate_password_hash(self.pswd1.data)
						wb.save('sheet.xlsx')
						return True
		else:
			return False

def token_required(func):
	@wraps(func)
	def decorated(*args, **kwargs):
		token = request.args.get('token')
		if not token:
			return 'token is missing'
		try:
			data = jwt.decode(token, app.config['SECRET_KEY'], algorithms=["HS256"])
		except:
			return 'Invalid token'
		return func(*args, **kwargs)
	return decorated

@app.route('/', methods=['GET','POST'])
def login():
	uname = None
	token = {}
	form = LoginForm()
	if form.validate_on_submit():
		uname = form.uname.data
		form.uname.data = ''
		form.pswd.data = ''
		token = jwt.encode({
				'user': uname,
				'exp': datetime.utcnow() + timedelta(seconds=100)
		},
				app.config['SECRET_KEY'])
	else:
		uname = 'error'
	return render_template('index.html', uname = uname, form = form, token = token)

@app.route('/change_pswd/<usr>', methods=['GET','POST'])
@token_required
def chng_pswd(usr):
	form = PswdForm()
	pswd1 = None
	pswd2 = None
	if form.change_pswd(usr):
		pswd1 = 'set successfully'
		form.pswd1.data = ''
		form.pswd2.data = ''
	return render_template('/pswd_chng.html', status = pswd1, form = form)


@app.route('/<command>/<usr>', methods=['GET','POST'])
@token_required
def process(command, usr):
	if command == 'open':
		logs['A'+str(logs.max_row+1)] = usr + ' opened door at ' + str(datetime.utcnow()) + ' (UTC)' 
		servo.value = -1
	elif command == 'close':
		logs['A'+str(logs.max_row+1)] = usr + ' closed door at ' + str(datetime.utcnow()) + ' (UTC)' 
		servo.value = 1
	wb.save('sheet.xlsx')
	return render_template('thankyou.html', cmd = command)

	
if __name__ == '__main__':
	app.run(debug=True, host ='0.0.0.0')